import openpyxl

def getRowCount(file, sheetName) -> int:
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file, sheetName) -> int:
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)

def readData(file, sheetName, rowNum, colNum) -> str:
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    data = sheet.cell(row=rowNum, column=colNum).value
    return "" if data == None else data

def writeData(file, sheetName, rowNum, colNum, data) -> None:
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value = data
    workbook.save(file)

# row_count = writeData("./login_data.xlsx", "loginSecurity", 1, 1, "test")
# print(row_count)